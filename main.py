from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from dotenv import load_dotenv, set_key
import getpass
import os
import time
from datetime import datetime

def calenderScrape():
    service = Service()
    option = webdriver.ChromeOptions()
    option.add_argument("--headless")
    driver = webdriver.Chrome(service=service, options=option)
    load_dotenv("info.env")
    
    url = "https://estudijas.rtu.lv/"
    driver.get(url)
    driver.find_element(By.XPATH, "/html/body/div[3]/header[1]/div/div/div[2]/div/form/div[1]/button").click()
    elem = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td[2]/form/input")
    elem.send_keys(os.getenv("USER"))
    elem = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/table/tbody/tr[2]/td[2]/form/input")
    elem.send_keys(os.getenv("PASS"))
    driver.find_element(By.XPATH, "/html/body/div[3]/div/div[1]/table/tbody/tr[3]/td[2]/input").click()
    time.sleep(1)
    if driver.title == "ORTUS (Neveiksmīga autentifikācija)":
        driver.quit()
        updateKeys()
    
    eventNames = []
    eventTimes = []
    events = driver.find_elements(By.CLASS_NAME, "overflow-auto")

    for val in events:
        val = val.text
        val = val.splitlines()
        try:
            eventNames.append(val[0])
            eventTimes.append(val[1])
        except IndexError:
            pass
    print("Event list scraped.")
    tableGen(eventNames, eventTimes)

def tableGen(eventNames, eventTimes):
    wb = Workbook()
    ws = wb.active
    
    ws["A1"]="Event name"
    ws["B1"]="End date"
    ws["C1"]="Time left"
    row=2
    i=0
    for val in eventNames:
        date = calculateDate(eventTimes[i])
        ws[f"A{row}"] = val
        ws[f"B{row}"] = date
        if date > datetime.now():
            ws[f"C{row}"] = f"=B{row}-NOW()"
        else: 
            ws[f"C{row}"] = "Past due"
        row += 1
        i += 1
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row+1, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY h:mm'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row+1, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = 'd "days" h "hours and" mm "minutes"'
    
    columnWidths = {'A': 50, 'B': 20, 'C': 30}
    for column, width in columnWidths.items():
        ws.column_dimensions[column].width = width

    wb.save("planned_events.xlsx")
    print("Excel spreadsheet created.")

def calculateDate(eventTime):
    latvianMonths = {
    'janvāris': 1,
    'februāris': 2,
    'marts': 3,
    'aprīlis': 4,
    'maijs': 5,
    'jūnijs': 6,
    'jūlijs': 7,
    'augusts': 8,
    'septembris': 9,
    'oktobris': 10,
    'novembris': 11,
    'decembris': 12
    }
    eventTimeTest = eventTime.split(",")
    if eventTimeTest[0] != "Šodien":
        eventTime = eventTime.split(", ")
        eventTime = f"{eventTime[1]}, {eventTime[2]}"
    datePart, timePart = eventTime.split(", ")
    if datePart == "Šodien":
        day = datetime.now().day
        month = datetime.now().month
    else:
        day, monthLatvian = datePart.split('. ')
        month = latvianMonths[monthLatvian.lower()]
    year = datetime.now().year

    dateString = f"{int(day):02d}.{month:02d}.{year} {timePart}"
    return datetime.strptime(dateString, "%d.%m.%Y %H:%M")

def updateKeys():
    print("Invalid login info. Update .env file keys.")
    os.environ["USER"] = input("Enter new ORTUS username: ")
    os.environ["PASS"] = getpass.getpass("Enter new ORTUS password: ")
    set_key("info.env", "USER", os.environ["USER"])
    set_key("info.env", "PASS", os.environ["PASS"])
    print("Keys updated.")
    calenderScrape()

if __name__ == "__main__":
    calenderScrape()
    print("Program complete. You can close the program.")